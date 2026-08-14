#!/usr/bin/env python3
"""Builds the shareable proposal workbook: assets/reliance-proposal-deliverables-schedule.xlsx

The workbook is the Excel counterpart to the DOCX/PDF proposal exports. Those are
read-only artifacts for signature; this one is the working copy a client or a
teammate opens, edits, and sends back — line items in or out, quantities changed,
comments in the margin.

Its price book and its arithmetic are BOTH derived from the platform rather than
retyped, so an edited workbook and a generated proposal quote the same numbers:

  * the Price Book sheet is read out of lib/proposals/catalog.ts (via node, so
    the TypeScript literals stay the one source of truth), and
  * the totals chain mirrors computeProposalTotals() in lib/proposals/pricing.ts:
        subtotal = SUM of rounded line amounts   (rows excluded with N score 0)
        discount = subtotal x discount%
        taxable  = subtotal - discount
        tax      = taxable  x tax%               (tax applies AFTER the discount)
        total    = taxable  + tax
        deposit  = total    x deposit%           (of total, not of subtotal)

Re-run it after any catalog price change:

    python3 scripts/build-proposal-workbook.py

Requires openpyxl (pip install openpyxl) and node on PATH. Recalculate the
result with LibreOffice afterwards so the formula cells carry cached values for
anything that reads the file without evaluating it (Google Sheets, previewers):

    soffice --headless --convert-to xlsx --outdir <dir> <file>
"""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

REPO = Path(__file__).resolve().parent.parent
CATALOG_TS = REPO / "lib" / "proposals" / "catalog.ts"
OUT = REPO / "assets" / "reliance-proposal-deliverables-schedule.xlsx"

# Brand palette, matching lib/proposals/docx.ts so the workbook prints as the
# same document family as the DOCX and PDF exports.
NAVY = "0C3450"
GOLD = "DCA23A"
INK = "16242F"
MUTED = "5D6F7D"
LINE = "DBE2E9"
BAND = "F1F6FA"
GOLD_TINT = "FBF2DD"
WHITE = "FFFFFF"

INPUT_BLUE = "0000FF"  # cells the reader types into
FILL_YELLOW = "FFFF00"  # cells that MUST be filled in before sending

FONT = "Arial"
MONEY = '$#,##0.00;($#,##0.00);"—"'
PCT = "0.0%"

thin = Side(style="thin", color=LINE)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)


def read_catalog() -> dict:
    """Evaluates the three object literals in catalog.ts without a TS toolchain."""
    script = r"""
    const { readFileSync } = require("node:fs");
    const src = readFileSync(process.argv[1], "utf8");
    function grab(name) {
      const start = src.indexOf(`export const ${name} = freezeCatalog({`);
      if (start === -1) throw new Error(`missing ${name}`);
      const open = src.indexOf("{", start + `export const ${name} = freezeCatalog(`.length - 1);
      let depth = 0, quote = null, i = open;
      for (; i < src.length; i++) {
        const ch = src[i];
        if (quote) { if (ch === "\\") { i++; continue; } if (ch === quote) quote = null; continue; }
        if (ch === '"' || ch === "'" || ch === "`") { quote = ch; continue; }
        if (ch === "{") depth++;
        else if (ch === "}") { depth--; if (depth === 0) break; }
      }
      return new Function(`return ${src.slice(open, i + 1)};`)();
    }
    process.stdout.write(JSON.stringify({
      phaseOptions: grab("phaseOptions"),
      serviceOptions: grab("serviceOptions"),
      packageData: grab("packageData"),
    }));
    """
    proc = subprocess.run(
        ["node", "-e", script, str(CATALOG_TS)],
        capture_output=True,
        text=True,
        check=True,
    )
    return json.loads(proc.stdout)


# Display order of the service picker's option groups (catalog.ts: serviceGroups).
SERVICE_GROUPS = [
    "Platform & Licensing",
    "Implementation & Consulting",
    "Safety Documents & Programs",
    "Training Catalog",
    "Audits & Field Support",
    "Travel & Expenses",
    "Custom",
]


def build_price_book(catalog: dict) -> list[dict]:
    """One flat lookup table: every quotable line, in document order."""
    rows: list[dict] = []
    for _key, pkg in catalog["packageData"].items():
        if pkg["price"] == 0 and pkg["users"] == 0 and pkg["sites"] == 0:
            # "Platform Services" / "Services Engagement" carry no price of their
            # own; they exist so a services-only deal can omit the package row.
            continue
        rows.append(
            {
                "name": pkg["name"],
                "type": "Package",
                "group": "Platform & Licensing",
                "unit": "Term",
                "price": pkg["price"],
                "desc": f"{pkg['desc']} Includes {pkg['users']} users across {pkg['sites']} jobsites.",
            }
        )
    for _key, phase in catalog["phaseOptions"].items():
        if phase["price"] == 0:
            continue  # "Custom Phase" — priced per proposal, added by hand.
        rows.append(
            {
                "name": phase["name"],
                "type": "Phase",
                "group": "Implementation Phases",
                "unit": "Phase",
                "price": phase["price"],
                "desc": phase["desc"],
            }
        )
    services = list(catalog["serviceOptions"].values())
    for group in SERVICE_GROUPS:
        for svc in services:
            if svc["group"] != group or svc["price"] == 0:
                continue  # zero-priced entries are "custom scope, set your own".
            rows.append(
                {
                    "name": svc["name"],
                    "type": "Service",
                    "group": svc["group"],
                    "unit": svc["unit"],
                    "price": svc["price"],
                    "desc": svc["desc"],
                }
            )

    seen: dict[str, int] = {}
    for row in rows:
        if row["name"] in seen:
            raise SystemExit(
                f"Duplicate catalog name {row['name']!r}: the workbook looks lines up by name."
            )
        seen[row["name"]] = 1
    return rows


# The example proposal the workbook opens with — a mid-size platform rollout.
# Every name here must exist in the price book; build() asserts it.
EXAMPLE_LINES = [
    ("Y", "Professional Safety Intelligence", 1),
    ("Y", "Discovery & Intake", 1),
    ("Y", "Build & Configure", 1),
    ("Y", "Validation & Review", 1),
    ("Y", "Launch & Training", 1),
    ("Y", "Predictive Risk Intelligence", 1),
    ("Y", "Training Matrix & Certification Tracking", 1),
    ("Y", "Mobile Field App Enablement", 1),
    ("N", "AI Gateway Validation Layer", 1),
    ("Y", "Safety Document — Medium (≤60 pg)", 4),
    ("Y", "OSHA 30 Training", 6),
    ("Y", "Compliance Audit", 2),
    ("Y", "Field Support Day", 4),
    ("Y", "Travel Mileage", 850),
    ("Y", "Hotel Night", 6),
    ("Y", "Per Diem", 6),
]

# Deliverable rows seeded from the implementation phases, with the acceptance
# criteria language the fixed-price proposal profile uses.
EXAMPLE_DELIVERABLES = [
    (
        "Discovery & Intake",
        "Discovery findings and configuration plan",
        "Objectives workshop, current document inventory, user and role map, jobsite list, and the agreed configuration priorities.",
        "Client confirms in writing that the user map, jobsite list and configuration priorities are correct.",
        "Joint",
    ),
    (
        "Build & Configure",
        "Configured platform tenant",
        "Modules, templates, dashboards, workflows, permission groups and billing package selections configured in the client tenant.",
        "Every module in the schedule is reachable by its intended role in the client tenant.",
        "Reliance",
    ),
    (
        "Validation & Review",
        "Validation report and corrected configuration",
        "Sample outputs reviewed, required fields confirmed, reporting logic verified, workflows tested, gaps corrected before launch.",
        "All defects found in validation are closed or accepted in writing by the client.",
        "Joint",
    ),
    (
        "Launch & Training",
        "Go-live and trained user base",
        "Rollout support, user training sessions, manager review, reporting cadence, and go-live stabilization.",
        "Named users complete training and the platform is in production use at the agreed sites.",
        "Reliance",
    ),
    (
        "Predictive Risk Intelligence",
        "Leading indicator model and executive dashboard",
        "Precursor cells, trend logic, risk scoring, dashboard review and the executive summary output.",
        "Risk scores render for each configured site and the executive summary reconciles to source records.",
        "Reliance",
    ),
    (
        "Training Matrix & Certification Tracking",
        "Role-based training matrix",
        "Training matrix by role, expiration tracking, missing-training summaries and compliance reporting.",
        "Every role in the client's org chart maps to a training requirement set with expiry dates.",
        "Reliance",
    ),
    (
        "Mobile Field App Enablement",
        "Field app in production",
        "Field issue capture, JSA support, inspection entry, photo observations and mobile user rollout support.",
        "Field users can submit an inspection and a photo observation from a jobsite device.",
        "Reliance",
    ),
    (
        "Safety Document — Medium (≤60 pg)",
        "Four written safety programs",
        "Four safety programs drafted to the client's operations, each up to 60 pages, with a revision record.",
        "Each program is delivered in the agreed format and cites the regulations applicable to the client's scopes of work.",
        "Reliance",
    ),
    (
        "OSHA 30 Training",
        "Six supervisors OSHA 30 certified",
        "OSHA 30-hour outreach course for supervisors and safety leads. Department of Labor cards issued to each attendee.",
        "Six named attendees complete the course and receive DOL cards.",
        "Reliance",
    ),
    (
        "Compliance Audit",
        "Two compliance audits with corrective action plans",
        "Structured audit against OSHA and company program requirements, delivered as a scored findings report.",
        "Each audit is delivered with prioritized corrective actions and due dates assigned to named owners.",
        "Reliance",
    ),
    (
        "Field Support Day",
        "Four field support days",
        "Pre-task briefings, field observations, corrective coaching and a written end-of-day summary for management.",
        "A written end-of-day summary is delivered to management for each day worked.",
        "Reliance",
    ),
    (
        "Client obligations",
        "Client-side inputs",
        "Current safety documents, org chart and user list, site list, and a named point of contact for approvals.",
        "Inputs are provided within 10 business days of kickoff; delays move the schedule day for day.",
        "Client",
    ),
]


def style_title(ws, row: int, text: str, sub: str, width: int) -> None:
    ws.cell(row=row, column=1, value=text).font = Font(
        name=FONT, size=16, bold=True, color=NAVY
    )
    ws.cell(row=row + 1, column=1, value=sub).font = Font(
        name=FONT, size=9, italic=True, color=MUTED
    )
    for col in range(1, width + 1):
        ws.cell(row=row + 2, column=col).fill = PatternFill("solid", fgColor=GOLD)
    ws.row_dimensions[row + 2].height = 3


def header_row(ws, row: int, headers: list[str]) -> None:
    for idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=text)
        cell.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BOX
    ws.row_dimensions[row].height = 30


def label(ws, row: int, col: int, text: str, *, bold: bool = False, color: str = INK) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name=FONT, size=10, bold=bold, color=color)


def input_cell(ws, row: int, col: int, value=None, *, fmt: str | None = None, required: bool = False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name=FONT, size=10, color=INPUT_BLUE)
    cell.border = BOX
    if fmt:
        cell.number_format = fmt
    if required:
        cell.fill = PatternFill("solid", fgColor=FILL_YELLOW)
    return cell


def formula_cell(ws, row: int, col: int, formula: str, *, fmt: str | None = None, bold: bool = False):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = Font(name=FONT, size=10, bold=bold, color=INK)
    cell.border = BOX
    if fmt:
        cell.number_format = fmt
    return cell


def build() -> None:
    catalog = read_catalog()
    book = build_price_book(catalog)
    names = {row["name"] for row in book}
    for _include, name, _qty in EXAMPLE_LINES:
        if name not in names:
            raise SystemExit(f"Example line {name!r} is not in the price book.")

    wb = Workbook()

    # ---------------------------------------------------------------- Price Book
    pb = wb.active
    pb.title = "Price Book"
    style_title(
        pb,
        1,
        "Price Book",
        "Reference only — do not edit. Generated from lib/proposals/catalog.ts; "
        "the Fee Schedule looks every line up here by name.",
        6,
    )
    PB_HEADER = 5
    header_row(pb, PB_HEADER, ["Line Item", "Type", "Group", "Unit", "List Price (USD)", "Standard Scope Description"])
    for offset, row in enumerate(book):
        r = PB_HEADER + 1 + offset
        pb.cell(row=r, column=1, value=row["name"]).font = Font(name=FONT, size=9, bold=True, color=INK)
        pb.cell(row=r, column=2, value=row["type"]).font = Font(name=FONT, size=9, color=MUTED)
        pb.cell(row=r, column=3, value=row["group"]).font = Font(name=FONT, size=9, color=MUTED)
        pb.cell(row=r, column=4, value=row["unit"]).font = Font(name=FONT, size=9, color=MUTED)
        price = pb.cell(row=r, column=5, value=row["price"])
        price.font = Font(name=FONT, size=9, color=INK)
        price.number_format = MONEY
        desc = pb.cell(row=r, column=6, value=row["desc"])
        desc.font = Font(name=FONT, size=9, color=INK)
        desc.alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(1, 7):
            pb.cell(row=r, column=col).border = BOX
            if offset % 2 == 1:
                pb.cell(row=r, column=col).fill = PatternFill("solid", fgColor=BAND)
    pb_first = PB_HEADER + 1
    pb_last = PB_HEADER + len(book)
    for col, width in zip("ABCDEF", [46, 11, 28, 11, 16, 96]):
        pb.column_dimensions[col].width = width
    pb.freeze_panes = f"A{pb_first}"

    def pb_range(col: str) -> str:
        return f"'Price Book'!${col}${pb_first}:${col}${pb_last}"

    for name, col in [
        ("CatalogNames", "A"),
        ("CatalogType", "B"),
        ("CatalogGroup", "C"),
        ("CatalogUnit", "D"),
        ("CatalogPrice", "E"),
        ("CatalogDesc", "F"),
    ]:
        wb.defined_names.add(DefinedName(name, attr_text=pb_range(col)))

    # -------------------------------------------------------------- Fee Schedule
    fs = wb.create_sheet("Fee Schedule")
    style_title(
        fs,
        1,
        "Schedule of Fees",
        "EDIT THIS SHEET. Blue cells are yours to change. Pick a line item from the "
        "dropdown and the unit, list price and scope fill themselves in.",
        9,
    )
    FS_HEADER = 5
    header_row(
        fs,
        FS_HEADER,
        [
            "Include?\n(Y / N)",
            "Group",
            "Line Item\n(pick from list)",
            "Scope / Description\n(type over to tailor)",
            "Unit",
            "Qty",
            "List Price",
            "Agreed Price\n(blank = list)",
            "Line Total",
        ],
    )
    FS_FIRST = FS_HEADER + 1
    FS_ROWS = 45
    FS_LAST = FS_FIRST + FS_ROWS - 1

    for idx in range(FS_ROWS):
        r = FS_FIRST + idx
        seed = EXAMPLE_LINES[idx] if idx < len(EXAMPLE_LINES) else None

        inc = input_cell(fs, r, 1, seed[0] if seed else None)
        inc.alignment = Alignment(horizontal="center")

        formula_cell(fs, r, 2, f'=IF($C{r}="","",IFERROR(INDEX(CatalogGroup,MATCH($C{r},CatalogNames,0)),"Custom"))').font = Font(
            name=FONT, size=9, color=MUTED
        )
        item = input_cell(fs, r, 3, seed[1] if seed else None)
        item.alignment = Alignment(wrap_text=True, vertical="top")

        desc = formula_cell(
            fs,
            r,
            4,
            f'=IF($C{r}="","",IFERROR(INDEX(CatalogDesc,MATCH($C{r},CatalogNames,0)),""))',
        )
        desc.font = Font(name=FONT, size=9, color=INK)
        desc.alignment = Alignment(wrap_text=True, vertical="top")

        unit = formula_cell(fs, r, 5, f'=IF($C{r}="","",IFERROR(INDEX(CatalogUnit,MATCH($C{r},CatalogNames,0)),"Unit"))')
        unit.font = Font(name=FONT, size=9, color=MUTED)
        unit.alignment = Alignment(horizontal="center")

        qty = input_cell(fs, r, 6, seed[2] if seed else None, fmt="#,##0.##")
        qty.alignment = Alignment(horizontal="center")

        formula_cell(
            fs,
            r,
            7,
            f'=IF($C{r}="","",IFERROR(INDEX(CatalogPrice,MATCH($C{r},CatalogNames,0)),0))',
            fmt=MONEY,
        )
        input_cell(fs, r, 8, fmt=MONEY)
        formula_cell(
            fs,
            r,
            9,
            f'=IF($C{r}="","",IF(UPPER($A{r})="N",0,ROUND($F{r}*IF($H{r}="",$G{r},$H{r}),2)))',
            fmt=MONEY,
            bold=True,
        )
        fs.row_dimensions[r].height = 34

    subtotal_row = FS_LAST + 2
    label(fs, subtotal_row, 7, "Subtotal", bold=True)
    fs.cell(row=subtotal_row, column=7).alignment = Alignment(horizontal="right")
    sub = formula_cell(fs, subtotal_row, 9, f"=ROUND(SUM($I${FS_FIRST}:$I${FS_LAST}),2)", fmt=MONEY, bold=True)
    sub.fill = PatternFill("solid", fgColor=GOLD_TINT)
    sub.comment = Comment(
        "Sum of the ROUNDED line amounts, matching computeProposalTotals() in "
        "lib/proposals/pricing.ts — the client adds up the printed rows, so the "
        "subtotal has to agree with them to the cent.",
        "Reliance Predictive Safety Technologies",
    )
    note = fs.cell(
        row=subtotal_row + 2,
        column=1,
        value="Rows 6–21 are a worked example of a mid-size platform rollout. Replace them with your scope — "
        "clear the Line Item cell to blank a row, or set Include? to N to keep it visible at $0 for the client to consider.",
    )
    note.font = Font(name=FONT, size=9, italic=True, color=MUTED)

    yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True, showDropDown=False)
    yn.error = "Enter Y to include this line in the total, or N to price it at zero."
    fs.add_data_validation(yn)
    yn.add(f"A{FS_FIRST}:A{FS_LAST}")

    items = DataValidation(type="list", formula1="=CatalogNames", allow_blank=True, showDropDown=False)
    items.error = "Pick a line from the Price Book, or type a custom name and set the Agreed Price yourself."
    items.errorStyle = "warning"
    items.promptTitle = "Line item"
    items.prompt = "Pick from the price book, or type your own and fill in Agreed Price."
    fs.add_data_validation(items)
    items.add(f"C{FS_FIRST}:C{FS_LAST}")

    for col, width in zip("ABCDEFGHI", [10, 24, 40, 74, 10, 8, 14, 15, 16]):
        fs.column_dimensions[col].width = width
    fs.freeze_panes = f"A{FS_FIRST}"
    fs.sheet_properties.pageSetUpPr.fitToPage = True
    fs.page_setup.orientation = "landscape"
    fs.page_setup.fitToWidth = 1
    fs.page_setup.fitToHeight = 0

    # ---------------------------------------------------------- Proposal Summary
    ps = wb.create_sheet("Proposal Summary", 0)
    style_title(
        ps,
        1,
        "Proposal — Deliverables & Fee Schedule",
        "Reliance Predictive Safety Technologies · working copy for review and markup",
        4,
    )

    fields = [
        ("Proposal number", None, True),
        ("Proposal date", None, True),
        ("Prepared by", None, True),
        ("Valid until", None, True),
        ("", None, False),
        ("Client company", None, True),
        ("Client address", None, False),
        ("Client contact", None, False),
        ("Contact email", None, False),
        ("", None, False),
        ("Engagement term start", None, False),
        ("Engagement term end", None, False),
        ("Term (months)", 12, False),
        ("Included users", 50, False),
        ("Included jobsites", 5, False),
    ]
    row = 5
    label(ps, row, 1, "PROPOSAL DETAILS", bold=True, color=NAVY)
    row += 1
    for text, default, required in fields:
        if text == "":
            row += 1
            continue
        label(ps, row, 1, text)
        input_cell(ps, row, 2, default, required=required)
        row += 1

    money_row = row + 1
    label(ps, money_row, 1, "COMMERCIAL TERMS", bold=True, color=NAVY)
    money_row += 1

    label(ps, money_row, 1, "Discount %")
    disc_pct = input_cell(ps, money_row, 2, 0.0, fmt=PCT)
    disc_pct.comment = Comment(
        "Applied to the subtotal. Tax is charged on the discounted figure, "
        "matching the platform's proposal math.",
        "Reliance Predictive Safety Technologies",
    )
    tax_row = money_row + 1
    label(ps, tax_row, 1, "Tax %")
    input_cell(ps, tax_row, 2, 0.0, fmt=PCT)
    dep_row = tax_row + 1
    label(ps, dep_row, 1, "Deposit % due at acceptance")
    dep = input_cell(ps, dep_row, 2, 0.25, fmt=PCT)
    dep.comment = Comment(
        "A percentage of the TOTAL, not of the subtotal.",
        "Reliance Predictive Safety Technologies",
    )

    tot_row = dep_row + 2
    label(ps, tot_row, 1, "TOTALS", bold=True, color=NAVY)
    tot_row += 1
    totals = [
        ("Subtotal", f"='Fee Schedule'!$I${subtotal_row}"),
        ("Discount", f"=-ROUND($B${tot_row}*$B${money_row},2)"),
        ("Taxable amount", f"=$B${tot_row}+$B${tot_row + 1}"),
        ("Tax", f"=ROUND($B${tot_row + 2}*$B${tax_row},2)"),
        ("Total", f"=$B${tot_row + 2}+$B${tot_row + 3}"),
        ("Deposit due at acceptance", f"=ROUND($B${tot_row + 4}*$B${dep_row},2)"),
        ("Balance on completion", f"=$B${tot_row + 4}-$B${tot_row + 5}"),
    ]
    for offset, (text, formula) in enumerate(totals):
        r = tot_row + offset
        is_total = text == "Total"
        label(ps, r, 1, text, bold=is_total)
        cell = formula_cell(ps, r, 2, formula, fmt=MONEY, bold=is_total)
        if is_total:
            cell.fill = PatternFill("solid", fgColor=GOLD_TINT)
            cell.font = Font(name=FONT, size=12, bold=True, color=NAVY)

    legend_row = tot_row + len(totals) + 2
    label(ps, legend_row, 1, "HOW TO READ THIS FILE", bold=True, color=NAVY)
    legend = [
        "Blue text = type here. Black text = calculated, leave it alone.",
        "Yellow cells must be filled in before this goes to a client.",
        "Fee Schedule is where the money is set. Deliverables is what we owe for it.",
        "Every figure on this sheet comes from the Fee Schedule — change a quantity there and the total follows.",
    ]
    for offset, text in enumerate(legend):
        cell = ps.cell(row=legend_row + 1 + offset, column=1, value=f"·  {text}")
        cell.font = Font(name=FONT, size=9, color=MUTED)

    ps.column_dimensions["A"].width = 34
    ps.column_dimensions["B"].width = 42
    ps.column_dimensions["C"].width = 4

    # --------------------------------------------------------------- Deliverables
    dl = wb.create_sheet("Deliverables", 2)
    style_title(
        dl,
        1,
        "Deliverables",
        "EDIT THIS SHEET. One row per thing we owe. Acceptance criteria are what the client "
        "measures it against — argue them here, not after invoicing.",
        9,
    )
    DL_HEADER = 5
    header_row(
        dl,
        DL_HEADER,
        [
            "#",
            "Fee Schedule line",
            "Deliverable",
            "What it includes",
            "Acceptance criteria",
            "Owner",
            "Target date",
            "Status",
            "Client comments",
        ],
    )
    DL_FIRST = DL_HEADER + 1
    DL_ROWS = 30
    DL_LAST = DL_FIRST + DL_ROWS - 1
    for idx in range(DL_ROWS):
        r = DL_FIRST + idx
        seed = EXAMPLE_DELIVERABLES[idx] if idx < len(EXAMPLE_DELIVERABLES) else None
        num = formula_cell(dl, r, 1, f'=IF($C{r}="","",COUNTA($C${DL_FIRST}:$C{r}))')
        num.alignment = Alignment(horizontal="center")
        if seed:
            input_cell(dl, r, 2, seed[0])
            input_cell(dl, r, 3, seed[1])
            input_cell(dl, r, 4, seed[2])
            input_cell(dl, r, 5, seed[3])
            input_cell(dl, r, 6, seed[4])
            input_cell(dl, r, 8, "Proposed")
        else:
            for col in (2, 3, 4, 5, 6, 8):
                input_cell(dl, r, col)
        input_cell(dl, r, 7)
        input_cell(dl, r, 9)
        for col in (2, 3, 4, 5, 9):
            dl.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        for col in (6, 7, 8):
            dl.cell(row=r, column=col).alignment = Alignment(horizontal="center", vertical="top")
        dl.row_dimensions[r].height = 46
        if idx % 2 == 1:
            for col in range(1, 10):
                dl.cell(row=r, column=col).fill = PatternFill("solid", fgColor=BAND)

    owner_dv = DataValidation(type="list", formula1='"Reliance,Client,Joint"', allow_blank=True, showDropDown=False)
    dl.add_data_validation(owner_dv)
    owner_dv.add(f"F{DL_FIRST}:F{DL_LAST}")
    status_dv = DataValidation(
        type="list", formula1='"Proposed,Agreed,Changed,Removed,Deferred"', allow_blank=True, showDropDown=False
    )
    dl.add_data_validation(status_dv)
    status_dv.add(f"H{DL_FIRST}:H{DL_LAST}")
    line_dv = DataValidation(type="list", formula1="=CatalogNames", allow_blank=True, showDropDown=False)
    line_dv.errorStyle = "warning"
    dl.add_data_validation(line_dv)
    line_dv.add(f"B{DL_FIRST}:B{DL_LAST}")

    count_row = DL_LAST + 2
    label(dl, count_row, 4, "Deliverables listed", bold=True)
    dl.cell(row=count_row, column=4).alignment = Alignment(horizontal="right")
    formula_cell(dl, count_row, 5, f'=COUNTA($C${DL_FIRST}:$C${DL_LAST})', bold=True)
    label(dl, count_row + 1, 4, "Still marked Proposed")
    dl.cell(row=count_row + 1, column=4).alignment = Alignment(horizontal="right")
    formula_cell(dl, count_row + 1, 5, f'=COUNTIF($H${DL_FIRST}:$H${DL_LAST},"Proposed")')

    for col, width in zip("ABCDEFGHI", [5, 34, 34, 60, 60, 11, 13, 13, 40]):
        dl.column_dimensions[col].width = width
    dl.freeze_panes = f"A{DL_FIRST}"
    dl.sheet_properties.pageSetUpPr.fitToPage = True
    dl.page_setup.orientation = "landscape"
    dl.page_setup.fitToWidth = 1
    dl.page_setup.fitToHeight = 0

    # ------------------------------------------------------------------ How to Use
    hu = wb.create_sheet("How to Use")
    style_title(
        hu,
        1,
        "How to use this workbook",
        "Read once, then work in Fee Schedule and Deliverables.",
        2,
    )
    guide = [
        ("What this is", ""),
        (
            "",
            "The working copy of a proposal — the version you send to a client or a colleague to mark up. "
            "The signed version is the PDF or Word file generated from the platform; this workbook is what "
            "you argue over first.",
        ),
        ("", ""),
        ("The four sheets", ""),
        ("Proposal Summary", "Who it is for, the commercial levers (discount, tax, deposit) and the totals. Totals are formulas."),
        ("Fee Schedule", "The money. One row per quoted line. This is the sheet that drives every total."),
        ("Deliverables", "What we owe for that money, and what the client measures it against at acceptance."),
        ("Price Book", "Reference. The full catalog with list prices. Do not edit — it is generated from the platform."),
        ("", ""),
        ("Colour code", ""),
        ("Blue text", "Type here. Quantities, agreed prices, dates, names, comments."),
        ("Black text", "Calculated. Overwriting one breaks the arithmetic for everything below it."),
        ("Yellow fill", "Must be filled in before this file goes to a client."),
        ("", ""),
        ("Editing the fee schedule", ""),
        ("Add a line", "Pick a Line Item from the dropdown. Unit, list price and scope text fill in automatically."),
        ("Discount a line", "Type your figure in Agreed Price. Blank means charge list price."),
        ("Drop a line", "Set Include? to N. The row stays visible for the client to see, priced at zero."),
        ("Quote something custom", "Type your own name in Line Item and set Agreed Price yourself. The lookup will not find it, which is fine."),
        ("Add more rows", "Insert rows inside the table, not below it, then copy the formulas down from the row above."),
        ("", ""),
        ("How the total is worked out", ""),
        ("", "subtotal = sum of the line totals"),
        ("", "discount = subtotal x discount %"),
        ("", "taxable  = subtotal - discount"),
        ("", "tax      = taxable x tax %   (tax is charged after the discount, not before)"),
        ("", "total    = taxable + tax"),
        ("", "deposit  = total x deposit %   (of the total, not of the subtotal)"),
        ("", ""),
        ("Assumptions baked in", ""),
        ("Prices", "List prices are the current platform catalog (lib/proposals/catalog.ts). They are defaults, not floors — override any of them in Agreed Price."),
        ("Example content", "The workbook ships with a worked mid-size platform rollout in rows 6-21 of Fee Schedule and the matching deliverables. It is an example. Replace it."),
        ("Currency", "USD throughout. Line totals are rounded to the cent per line, then added — same as the generated proposal."),
        ("Travel", "Mileage, hotel and per diem lines are estimates until the schedule is fixed."),
        ("", ""),
        ("Sending it back", ""),
        (
            "",
            "Mark deliverable rows Agreed / Changed / Removed, leave comments in the last column, and return the file. "
            "The agreed version is what gets loaded back into the platform and printed for signature.",
        ),
    ]
    r = 5
    for left, right in guide:
        if left and not right:
            cell = hu.cell(row=r, column=1, value=left.upper())
            cell.font = Font(name=FONT, size=11, bold=True, color=NAVY)
            r += 1
            continue
        if left:
            hu.cell(row=r, column=1, value=left).font = Font(name=FONT, size=10, bold=True, color=INK)
        if right:
            cell = hu.cell(row=r, column=2, value=right)
            cell.font = Font(name=FONT, size=10, color=INK)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    hu.column_dimensions["A"].width = 30
    hu.column_dimensions["B"].width = 108

    footer = hu.cell(
        row=r + 1,
        column=1,
        value="Reliance Predictive Safety Technologies · Proposal Form Rev. 15 · workbook generated from the platform price book",
    )
    footer.font = Font(name=FONT, size=8, italic=True, color=MUTED)

    # Reading order: the summary first, then the two sheets that get edited,
    # then the reference material.
    wb._sheets = [
        wb[name]
        for name in ["Proposal Summary", "Fee Schedule", "Deliverables", "Price Book", "How to Use"]
    ]
    wb.active = 0
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT.relative_to(REPO)} — {len(book)} catalog lines")


if __name__ == "__main__":
    sys.exit(build())
